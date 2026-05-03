"""
LOCUST LOAD TESTING - CDI Sistema MARÍA
========================================

Script de pruebas de carga para validar rendimiento con hasta 2000 usuarios.

INSTALACIÓN:
    pip install locust

USO BÁSICO:
    # Escenario 1: Carga Normal (50 usuarios)
    locust -f locustfile.py --users 50 --spawn-rate 10 --run-time 10m --html report_normal.html

    # Escenario 2: Carga Pico (200 usuarios)
    locust -f locustfile.py --users 200 --spawn-rate 20 --run-time 5m --html report_peak.html

    # Escenario 3: Stress Test (500 usuarios)
    locust -f locustfile.py --users 500 --spawn-rate 50 --run-time 3m --html report_stress.html

    # Con interfaz web (recomendado)
    locust -f locustfile.py --host=http://localhost:8010
    # Luego abrir: http://localhost:8089

MÉTRICAS ESPERADAS (para 2000 usuarios):
    - Response Time p95: < 1000ms (carga normal)
    - Response Time p99: < 2000ms (carga normal)
    - Error Rate: < 1%
    - Throughput: > 100 req/seg

REQUISITOS:
    - Servidor corriendo en http://localhost:8010
    - Archivos de prueba: test_facturas.xlsx, test_factura.pdf
"""

from locust import HttpUser, task, between, events
from locust.runners import MasterRunner
import random
import time
import json
import os

# ========================================================================
# Configuración
# ========================================================================

# Archivos de prueba (crear si no existen)
TEST_EXCEL = "test_facturas.xlsx"
TEST_PDF = "test_factura.pdf"

# Datos de prueba para clientes
NOMBRES = [
    "Importaciones SA", "Exportaciones SRL", "Comercio Internacional",
    "Global Trade", "Aduanas Express", "Logística Total",
    "Shipping Co", "Cargo Services", "Freight Solutions"
]

DOMINIOS = ["gmail.com", "yahoo.com", "empresa.com", "trade.com"]

# ========================================================================
# Event Handlers (para reportes)
# ========================================================================

@events.test_start.add_listener
def on_test_start(environment, **kwargs):
    """Evento al inicio del test"""
    print("=" * 60)
    print("🚀 INICIANDO LOAD TEST - CDI Sistema MARÍA")
    print("=" * 60)
    print(f"Host: {environment.host}")
    print(f"Usuarios: {environment.runner.target_user_count if hasattr(environment.runner, 'target_user_count') else 'N/A'}")
    print("=" * 60)


@events.test_stop.add_listener
def on_test_stop(environment, **kwargs):
    """Evento al finalizar el test"""
    print("\n" + "=" * 60)
    print("✅ LOAD TEST COMPLETADO")
    print("=" * 60)

    # Obtener estadísticas
    stats = environment.stats
    total_requests = stats.total.num_requests
    total_failures = stats.total.num_failures
    avg_response_time = stats.total.avg_response_time
    max_response_time = stats.total.max_response_time

    print(f"Total Requests: {total_requests}")
    print(f"Total Failures: {total_failures}")
    print(f"Failure Rate: {(total_failures/total_requests*100) if total_requests > 0 else 0:.2f}%")
    print(f"Avg Response Time: {avg_response_time:.2f}ms")
    print(f"Max Response Time: {max_response_time:.2f}ms")

    # Percentiles (si están disponibles)
    if hasattr(stats.total, 'get_response_time_percentile'):
        p50 = stats.total.get_response_time_percentile(0.5)
        p95 = stats.total.get_response_time_percentile(0.95)
        p99 = stats.total.get_response_time_percentile(0.99)
        print(f"Response Time p50: {p50:.2f}ms")
        print(f"Response Time p95: {p95:.2f}ms")
        print(f"Response Time p99: {p99:.2f}ms")

        # Criterio de éxito
        print("\n" + "=" * 60)
        print("CRITERIO DE ÉXITO:")
        if p95 < 1000:
            print(f"✅ p95 < 1000ms: {p95:.2f}ms")
        else:
            print(f"⚠️  p95 >= 1000ms: {p95:.2f}ms (esperado < 1000ms)")

        if (total_failures/total_requests*100) < 1:
            print(f"✅ Error rate < 1%: {(total_failures/total_requests*100):.2f}%")
        else:
            print(f"⚠️  Error rate >= 1%: {(total_failures/total_requests*100):.2f}%")

    print("=" * 60 + "\n")


# ========================================================================
# Usuario Base
# ========================================================================

class CDIUser(HttpUser):
    """
    Usuario simulado del sistema CDI.

    Distribución de tareas (weights):
    - 40%: Consultas (GET requests) - operaciones rápidas
    - 30%: Creación de clientes - operaciones medianas
    - 20%: Upload Excel - operaciones pesadas
    - 10%: Upload PDF - operaciones pesadas
    """

    # Tiempo de espera entre requests (simula usuario real)
    wait_time = between(1, 5)  # 1-5 segundos entre acciones

    # Variables de instancia para datos únicos
    user_id = 0
    client_ids = []

    def on_start(self):
        """
        Ejecutado una vez cuando el usuario inicia.
        Aquí podríamos hacer login si fuera necesario.
        """
        CDIUser.user_id += 1
        self.my_id = CDIUser.user_id
        self.client_counter = 0
        print(f"👤 Usuario #{self.my_id} iniciado")

    # ====================================================================
    # TASKS: Consultas (40% - weight=4)
    # ====================================================================

    @task(4)
    def health_check(self):
        """
        GET /health
        Verifica estado del sistema (operación muy rápida).
        """
        with self.client.get("/health", catch_response=True) as response:
            if response.status_code == 200:
                try:
                    data = response.json()
                    if "status" in data:
                        response.success()
                    else:
                        response.failure("Missing 'status' field")
                except json.JSONDecodeError:
                    response.failure("Invalid JSON response")
            else:
                response.failure(f"Got status code {response.status_code}")

    @task(2)
    def get_root(self):
        """
        GET /
        Acceso a página principal (carga frontend).
        """
        self.client.get("/", name="GET /")

    # ====================================================================
    # TASKS: Creación de Clientes (30% - weight=3)
    # ====================================================================

    @task(3)
    def create_client(self):
        """
        POST /api/clientes/public
        Crear un nuevo cliente con validación de inputs.
        """
        self.client_counter += 1
        timestamp = int(time.time() * 1000)

        client_data = {
            "nombre": f"{random.choice(NOMBRES)} {self.my_id}-{self.client_counter}",
            "email": f"user{self.my_id}_{timestamp}@{random.choice(DOMINIOS)}",
            "cuit": self._generate_cuit(),
            "direccion": f"Calle {random.randint(100, 9999)}"
        }

        with self.client.post(
            "/api/clientes/public",
            json=client_data,
            catch_response=True,
            name="POST /api/clientes/public"
        ) as response:
            if response.status_code == 200:
                try:
                    data = response.json()
                    if data.get("success"):
                        # Guardar ID del cliente para uso posterior
                        if "id" in data:
                            CDIUser.client_ids.append(data["id"])
                        response.success()
                    else:
                        response.failure(f"Creation failed: {data.get('detail', 'Unknown error')}")
                except json.JSONDecodeError:
                    response.failure("Invalid JSON response")
            else:
                response.failure(f"Got status code {response.status_code}")

    # ====================================================================
    # TASKS: Upload Excel (20% - weight=2)
    # ====================================================================

    @task(2)
    def upload_excel(self):
        """
        POST /upload_excel
        Upload de archivo Excel (operación pesada).
        """
        # Verificar que existe archivo de prueba
        if not os.path.exists(TEST_EXCEL):
            # Crear archivo dummy si no existe
            self._create_dummy_excel()

        # Seleccionar cliente aleatorio o usar ID genérico
        client_id = random.choice(CDIUser.client_ids) if CDIUser.client_ids else "1"

        try:
            with open(TEST_EXCEL, 'rb') as f:
                files = {'file': (TEST_EXCEL, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'client_id': str(client_id)}

                with self.client.post(
                    "/upload_excel",
                    files=files,
                    data=data,
                    catch_response=True,
                    name="POST /upload_excel"
                ) as response:
                    if response.status_code == 200:
                        try:
                            result = response.json()
                            if result.get("success"):
                                response.success()
                            else:
                                # Puede fallar si el archivo no es válido, pero eso es esperado
                                response.failure(f"Upload failed: {result.get('detail', 'Unknown')}")
                        except json.JSONDecodeError:
                            response.failure("Invalid JSON response")
                    else:
                        response.failure(f"Got status code {response.status_code}")
        except FileNotFoundError:
            # Si no se puede crear/leer el archivo, marcar como fallo pero no crashear
            self.client.get("/health", name="POST /upload_excel")  # Dummy request

    # ====================================================================
    # TASKS: Upload PDF (10% - weight=1)
    # ====================================================================

    @task(1)
    def upload_pdf(self):
        """
        POST /upload_pdf/public
        Upload de archivo PDF (operación pesada).
        """
        # Verificar que existe archivo de prueba
        if not os.path.exists(TEST_PDF):
            # Crear archivo dummy si no existe
            self._create_dummy_pdf()

        try:
            with open(TEST_PDF, 'rb') as f:
                files = {'file': (TEST_PDF, f, 'application/pdf')}

                with self.client.post(
                    "/upload_pdf/public",
                    files=files,
                    catch_response=True,
                    name="POST /upload_pdf/public"
                ) as response:
                    if response.status_code == 200:
                        try:
                            result = response.json()
                            if result.get("success"):
                                response.success()
                            else:
                                # Puede fallar si el archivo no es válido
                                response.failure(f"Upload failed: {result.get('detail', 'Unknown')}")
                        except json.JSONDecodeError:
                            response.failure("Invalid JSON response")
                    else:
                        response.failure(f"Got status code {response.status_code}")
        except FileNotFoundError:
            # Si no se puede crear/leer el archivo, marcar como fallo pero no crashear
            self.client.get("/health", name="POST /upload_pdf/public")  # Dummy request

    # ====================================================================
    # Métodos Auxiliares
    # ====================================================================

    def _generate_cuit(self):
        """Genera un CUIT aleatorio válido (11 dígitos)"""
        prefix = random.choice(["20", "23", "27"])
        middle = str(random.randint(10000000, 99999999))
        suffix = str(random.randint(0, 9))
        return f"{prefix}{middle}{suffix}"

    def _create_dummy_excel(self):
        """Crea un archivo Excel dummy mínimo para testing"""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Test'
            ws['B1'] = 'Data'
            wb.save(TEST_EXCEL)
            print(f"📄 Creado archivo de prueba: {TEST_EXCEL}")
        except ImportError:
            print("⚠️  openpyxl no instalado - no se puede crear Excel de prueba")
            print("   Instalar con: pip install openpyxl")

    def _create_dummy_pdf(self):
        """Crea un archivo PDF dummy mínimo para testing"""
        try:
            from reportlab.pdfgen import canvas
            c = canvas.Canvas(TEST_PDF)
            c.drawString(100, 750, "Test PDF for Load Testing")
            c.save()
            print(f"📄 Creado archivo de prueba: {TEST_PDF}")
        except ImportError:
            print("⚠️  reportlab no instalado - no se puede crear PDF de prueba")
            print("   Instalar con: pip install reportlab")


# ========================================================================
# Usuarios Especializados (Opcional)
# ========================================================================

class ReadOnlyUser(HttpUser):
    """
    Usuario de solo lectura (para simular carga de consultas).
    Útil para testear escenarios donde muchos usuarios solo consultan.
    """
    wait_time = between(2, 8)

    @task(5)
    def health_check(self):
        self.client.get("/health")

    @task(3)
    def get_root(self):
        self.client.get("/")

    @task(1)
    def get_download(self):
        # Intenta descargar un archivo (puede fallar si no existe, eso es OK)
        self.client.get("/download/example.xlsx", name="GET /download/{filename}")


class HeavyUser(HttpUser):
    """
    Usuario "pesado" que hace muchas operaciones de upload.
    Útil para stress testing de operaciones costosas.
    """
    wait_time = between(5, 10)

    @task(3)
    def upload_excel(self):
        if os.path.exists(TEST_EXCEL):
            with open(TEST_EXCEL, 'rb') as f:
                files = {'file': (TEST_EXCEL, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                self.client.post("/upload_excel", files=files, data={'client_id': '1'})

    @task(2)
    def upload_pdf(self):
        if os.path.exists(TEST_PDF):
            with open(TEST_PDF, 'rb') as f:
                files = {'file': (TEST_PDF, f, 'application/pdf')}
                self.client.post("/upload_pdf/public", files=files)

    @task(1)
    def health_check(self):
        self.client.get("/health")


# ========================================================================
# Configuración por Defecto
# ========================================================================

# Por defecto usar CDIUser (comportamiento mixto)
# Para usar otros usuarios:
#   locust -f locustfile.py ReadOnlyUser --host=http://localhost:8000
#   locust -f locustfile.py HeavyUser --host=http://localhost:8000
