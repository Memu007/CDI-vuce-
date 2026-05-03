"""
PDF Processing Router - Modular extraction endpoints for invoice PDFs

This module contains all PDF processing endpoints and helper functions
extracted from server_funcional.py to reduce its size and improve maintainability.

Endpoints:
- POST /process_operation/ - Process import operations
- POST /upload_pdf/ - Upload with robust pipeline
- POST /upload_pdf_llm/ - Upload with Gemini ALWAYS architecture
- POST /upload_pdf_gemini_only/ - Gemini only, no fallbacks

Helper functions:
- _extract_pdf_text() - Extract text from PDF
- _robust_extract_pdf_items() - Robust extraction pipeline
- _parse_pdf_tables() - Parse tables from PDF
- _llm_extract_pdf_items() - LLM extraction
- _fallback_extraction() - Fallback when LLM fails
- _evaluate_extraction_quality() - Quality assessment
- Multiple format-specific extractors
"""

from fastapi import APIRouter, UploadFile, File, Depends, Request, Body, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from typing import Dict, List
import os
import sys
import re
import math
import io
import tempfile
import subprocess
import glob
from datetime import datetime

# Project imports
try:
    from proyecto_maria.auth import require_role
    from proyecto_maria.models.operations import OperationPayload
    from proyecto_maria.core.logging_config import get_logger
    from proyecto_maria.services.monitoring_service import monitoring_service
    from proyecto_maria.services.cache_service import llm_cache
    from pydantic import ValidationError
except ImportError:
    from ..auth import require_role
    from ..models.operations import OperationPayload
    from ..core.logging_config import get_logger
    from ..services.monitoring_service import monitoring_service
    try:
        from ..services.cache_service import llm_cache
    except ImportError:
        llm_cache = None
    from pydantic import ValidationError

# Check if infrastructure is available
try:
    from proyecto_maria.services.monitoring_service import monitoring_service
    INFRASTRUCTURE_AVAILABLE = True
except ImportError:
    INFRASTRUCTURE_AVAILABLE = False

# Security modules (Blue Team hardening)
try:
    from proyecto_maria.security.file_security import validate_file_upload
    from proyecto_maria.security.log_sanitizer import get_safe_error_message
    SECURITY_MODULES_AVAILABLE = True
except ImportError:
    print("⚠️ Security modules not available in pdf_router, using basic security only")
    SECURITY_MODULES_AVAILABLE = False

# Setup logger
logger = get_logger("maria.pdf_router")
app_logger = get_logger("maria.app")

# Create router
router = APIRouter(tags=["PDF Processing"])

# Constants
DATA_DIR = os.getenv('DATA_DIR', 'data')


def _basic_pdf_security_checks(upload: UploadFile, data: bytes) -> None:
    """Simple validation when advanced modules are unavailable."""
    filename = (upload.filename or '').lower()
    if not filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF")

    content_type = (upload.content_type or '').lower()
    if content_type and 'pdf' not in content_type:
        raise HTTPException(status_code=400, detail="Tipo de archivo inválido (solo PDF)")

    sample = (data or b'')[:1024].lower()
    if sample.startswith(b"mz") or b"<script" in sample:
        raise HTTPException(status_code=400, detail="Contenido sospechoso detectado")

HEADER_SYNONYMS = {
    'descripcion': [
        'descripcion', 'descripción', 'descrição', 'description', 'descrição da mercadoria',
        'producto', 'product', 'item', 'goods', 'mercaderia', 'mercadería', 'commodity', 'articulo'
    ],
    'cantidad': [
        'cantidad', 'cant', 'qty', 'quantity', 'unidades', 'units', 'piezas', 'pieces', 'pcs',
        'quantidade', 'bultos', 'cantpiezas', 'cantidad piezas', 'quantidade peças', 'qty/cant',
        'cantidad kg', 'kg'
    ],
    'precio_unitario': [
        'precio unit', 'unit price', 'precio/price', 'us$/pc', 'usd/pc', 'precio unitario',
        'price', 'valor unit', 'valor unitario', 'p.u.', 'pu', 'moneda/unid', 'usd/kg', 'rate'
    ],
    'total': [
        'total', 'amount', 'importe', 'total por item', 'total usd', 'valor total', 'neto', 'net', 'subtotal', 'monto'
    ],
    'ncm': ['ncm', 'hs code', 'hsc', 'tariff', 'posicion', 'código', 'codigo', 'hs destino', 'commodity code', 'tarifa'],
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def _to_number_any(raw: object, default: float = 0.0) -> float:
    """Convert any raw value to float, handling various formats"""
    try:
        s = str(raw or '').strip()
        if not s:
            return default
        # Clean currency symbols and spaces
        s = s.replace('USD', '').replace('US$', '').replace('\xa0', ' ').strip()
        # Remove common currency abbreviations
        s = re.sub(r"(?i)(U\$S|USD|EUR|ARS|BRL|R\$|CLP|COP)", '', s)
        s = re.sub(r"[\$€£]", '', s)
        # If no digits
        if not re.search(r"\d", s):
            return default
        # Handle thousands and decimal separators
        if ',' in s and '.' in s:
            # If comma is to the right of dot, assume comma decimal (ES format)
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s and '.' not in s:
            # Only comma → assume decimal comma
            s = s.replace(',', '.')
        # Remove remaining spaces
        s = s.replace(' ', '')
        return float(s)
    except Exception:
        return default


def _to_float(val, default=0.0):
    """Convert value to float"""
    try:
        return float(str(val).replace(',', '.'))
    except Exception:
        return default


def _clean_ncm(val):
    """Clean and format NCM code"""
    s = ''.join(ch for ch in str(val) if ch.isdigit())
    if len(s) >= 4:
        return s[:8]  # Max 8 digits
    return s


def _detect_default_origin(text_rows: list[str]) -> str:
    """Detect default origin from text"""
    blob = " ".join(text_rows).lower()
    if 'argentina' in blob:
        return 'AR'
    if 'china' in blob:
        return 'CN'
    if 'vietnam' in blob or 'viet' in blob:
        return 'VN'
    if 'brazil' in blob or 'brasil' in blob:
        return 'BR'
    return 'XX'


def _tariff_group_from_pieza(pieza: str) -> str:
    """Extract tariff group from pieza (first 4 digits)"""
    if pieza and len(pieza) >= 4:
        return pieza[:4]
    return ''


def _is_noise_desc(desc: str) -> bool:
    """Check if description is noise (too short, only numbers, etc.)"""
    if not desc or len(desc) < 3:
        return True
    # Only numbers or special chars
    if re.match(r'^[\d\s\.,;:\-\+]+$', desc):
        return True
    # Common noise words
    noise_words = ['subtotal', 'total', 'tax', 'iva', 'impuesto', 'comments', 'observaciones', 'freight', 'page', 'página', 'fecha', 'date', 'invoice', 'factura']
    if any(word in desc.lower() for word in noise_words):
        return True
    return False


def _extract_pdf_text(data: bytes) -> str:
    """Extract text from PDF using pdfminer or PyPDF2 as fallback"""
    # Try pdfminer first (best accuracy), fallback to PyPDF2
    try:
        from pdfminer.high_level import extract_text
        return extract_text(io.BytesIO(data)) or ''
    except Exception:
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(data))
            pages = []
            for p in reader.pages:
                try:
                    pages.append(p.extract_text() or '')
                except Exception:
                    pass
            return '\n'.join(pages)
        except Exception:
            return ''


def _parse_pdf_tables(data: bytes) -> list[dict]:
    """Parse tables from PDF using pdfplumber if available"""
    try:
        import pdfplumber
        items: list[dict] = []
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                try:
                    # Try different table extraction strategies
                    table_settings_list = [
                        {"vertical_strategy": "lines", "horizontal_strategy": "lines", "snap_tolerance": 3, "join_tolerance": 3},
                        {"vertical_strategy": "text", "horizontal_strategy": "text"},
                    ]
                    tables = []
                    for ts in table_settings_list:
                        try:
                            tables = page.extract_tables(table_settings=ts) or []
                            if tables:
                                break
                        except Exception:
                            continue
                except Exception:
                    tables = []

                for tbl_i, tbl in enumerate(tables):
                    if not tbl or len(tbl) < 2:
                        continue

                    # Normalize header by first row with most text
                    header = None
                    for r in tbl[:3]:
                        if r and sum(1 for c in r if c and str(c).strip()) >= 3:
                            header = r
                            break
                    if not header:
                        header = tbl[0]
                    cols = [str(c or '').strip().lower() for c in header]

                    # Map columns
                    def _find(*keys):
                        for i, c in enumerate(cols):
                            if any(k in c for k in keys):
                                return i
                        return None

                    # Cover varied descriptions: Spanish/Portuguese/English
                    i_desc = _find('desc', 'descripción', 'descripcion', 'descrição', 'mercadería', 'mercaderia', 'product', 'detalle', 'item', 'goods')
                    # Quantity: prefer pieces/qty over kg
                    i_qty_pcs = _find('cantidad piezas', 'pcs', 'piezas', 'qty', 'qtd', 'qty/cant')
                    i_qty_kg = _find('cantidad kg', 'kg')
                    i_qty = i_qty_pcs if i_qty_pcs is not None else i_qty_kg
                    # Unit price
                    i_price = _find('unit price', 'unit', 'price', 'precio', 'usd/pc', 'us$/pc', 'valor', 'valor unit', 'usd', 'rate', 'p.u.', 'pu')
                    # NCM/HS code if exists
                    i_ncm = _find('hs code', 'hs-code', 'hs', 'ncm', 'tariff', 'posicion', 'posición', 'codigo', 'código')

                    # Requires at least desc + qty or price
                    if i_desc is None or (i_qty is None and i_price is None):
                        # Heuristic without reliable headers: use last numeric columns
                        def _row_to_item(r, row_no):
                            try:
                                if not r:
                                    return None
                                # Remove empty cells from tail
                                rr = list(r)
                                while rr and (rr[-1] is None or str(rr[-1]).strip() == ''):
                                    rr.pop()
                                if len(rr) < 4:
                                    return None
                                price = _to_number_any(rr[-2])
                                qty = _to_number_any(rr[-3])
                                if price <= 0 and qty <= 0:
                                    return None
                                desc = ' '.join(str(x or '').strip() for x in rr[:-3]).strip()
                                if len(desc) < 3 or _is_noise_desc(desc):
                                    return None
                                pieza_val = ''
                                return {
                                    'pieza': pieza_val,
                                    'descripcion': desc[:200],
                                    'origen': 'XX',
                                    'cantidad': qty if qty > 0 else 1.0,
                                    'valor_unitario': price if price >= 0 else 0.0,
                                    'peso_unitario': 0.0,
                                    'order_index': (tbl_i * 10000) + row_no,
                                    'tariff_group': _tariff_group_from_pieza(pieza_val),
                                }
                            except Exception:
                                return None

                        tmp_items = []
                        for ridx, r in enumerate(tbl[1:], start=1):
                            it = _row_to_item(r, ridx)
                            if it:
                                tmp_items.append(it)
                        if len(tmp_items) >= 2:
                            items.extend(tmp_items)
                            if len(items) >= 1:
                                break
                        continue

                    # Process data rows
                    for ridx, r in enumerate(tbl[1:], start=1):
                        if not r or len(r) <= max((i for i in [i_desc, i_qty, i_price] if i is not None), default=0):
                            continue
                        desc = str(r[i_desc] or '').strip()
                        if not desc or _is_noise_desc(desc):
                            continue
                        qty = _to_number_any(r[i_qty]) if i_qty is not None else 1.0
                        price = _to_number_any(r[i_price]) if i_price is not None else 0.0
                        if qty <= 0 and price <= 0:
                            continue
                        pieza_val = ''
                        try:
                            if i_ncm is not None and i_ncm < len(r):
                                pieza_val = _clean_ncm(r[i_ncm] or '')
                        except Exception:
                            pieza_val = ''
                        items.append({
                            'pieza': pieza_val,
                            'descripcion': desc[:200],
                            'origen': 'XX',
                            'cantidad': qty if qty > 0 else 1.0,
                            'valor_unitario': price if price >= 0 else 0.0,
                            'peso_unitario': 0.0,
                            'order_index': (tbl_i * 10000) + ridx,
                            'tariff_group': _tariff_group_from_pieza(pieza_val),
                        })
                        if len(items) >= 200:
                            break
                if len(items) >= 1:
                    break
        return items
    except Exception:
        return []


def _robust_extract_pdf_items(data: bytes) -> list[dict]:
    """
    Robust PDF extraction pipeline for invoice items.
    Returns list of item dicts with keys: pieza, descripcion, origen, cantidad, valor_unitario, peso_unitario, order_index, tariff_group
    """
    try:
        # Import the standalone module from project root
        root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
        if root_dir not in sys.path:
            sys.path.insert(0, root_dir)
        from pdf_extractor import robust_extract_pdf_items as standalone_extractor
        return standalone_extractor(data)
    except Exception:
        # Fallback to inline implementation
        pass

    # 1. Extract text
    text = _extract_pdf_text(data)
    if not text.strip():
        return []

    # 2. Split into lines and clean
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 5:
        return []

    # 3. Detect header row
    header_synonyms = HEADER_SYNONYMS
    header_idx = -1
    header_cols = {}

    for i, ln in enumerate(lines[:50]):  # Check first 50 lines
        low = ln.lower()
        score = 0
        cols_found = {}
        for field, synonyms in header_synonyms.items():
            for syn in synonyms:
                if syn.lower() in low:
                    cols_found[field] = syn.lower()
                    score += 1
                    break
        if score >= 2:  # At least 2 header fields
            header_idx = i
            header_cols = cols_found
            break

    if header_idx == -1:
        # Fallback: assume first line with numbers is header
        for i, ln in enumerate(lines[:20]):
            if re.search(r'\d', ln):
                header_idx = i
                break

    # 4. Parse rows after header
    items = []
    order_idx = 1
    current_desc = ''

    # Detect default origin
    default_origin = _detect_default_origin(lines)

    for i in range(header_idx + 1, len(lines)):
        ln = lines[i].strip()
        if not ln:
            continue

        low = ln.lower()

        # Skip noise lines
        noise_keywords = ['subtotal', 'tax', 'iva', 'impuesto', 'comments', 'observaciones', 'freight', 'grand total', 'total factura', 'total invoice', 'page', 'página', 'fecha', 'date', 'invoice', 'factura', 'n°', 'no.', 'number']
        if any(kw in low for kw in noise_keywords):
            continue

        # Extract numbers from line
        nums = re.findall(r'(?:\d{1,3}(?:[\.,]\d{3})*|\d+)(?:[\.,]\d{1,3})?', ln)
        if not nums:
            # Multiline description continuation
            if current_desc and not re.search(r'\d', ln):
                current_desc += ' ' + ln
            continue

        # Take last 2-4 numbers as numeric tail
        tail = nums[-4:] if len(nums) >= 4 else nums[-3:] if len(nums) >= 3 else nums[-2:] if len(nums) >= 2 else nums

        if len(tail) < 2:
            continue

        # Parse tail: [kg?, qty, price, total]
        vals = [_to_number_any(v) for v in tail]
        vals = [v for v in vals if v is not None]

        if len(vals) < 2:
            continue

        if len(tail) == 4:
            peso_unitario, cantidad, valor_unitario, total = vals[-4], vals[-3], vals[-2], vals[-1]
        elif len(tail) == 3:
            cantidad, valor_unitario, total = vals[-3], vals[-2], vals[-1]
            peso_unitario = 0.0
        else:
            cantidad, valor_unitario = vals[-2], vals[-1]
            total = 0.0
            peso_unitario = 0.0

        # If only total and qty, calculate unitario
        if cantidad > 0 and total > 0 and valor_unitario == 0:
            valor_unitario = total / cantidad

        if cantidad <= 0 or valor_unitario < 0:
            continue

        # Description: everything before the tail
        tail_start = ln.rfind(tail[0])
        desc = ln[:tail_start].strip() if tail_start > 0 else ln
        if current_desc:
            desc = current_desc + ' ' + desc
            current_desc = ''

        if len(desc) < 3:
            continue

        # Extract NCM: 6-8 digits
        ncm_match = re.search(r'\b(\d{6,8})\b', ln)
        pieza = ncm_match.group(1) if ncm_match else ''

        # If no NCM in line, check description
        if not pieza:
            ncm_match = re.search(r'\b(\d{6,8})\b', desc)
            pieza = ncm_match.group(1) if ncm_match else ''

        # Clean pieza to 8 digits max
        if pieza and len(pieza) > 8:
            pieza = pieza[:8]

        item = {
            'pieza': pieza or '',
            'descripcion': desc[:200],
            'origen': default_origin,
            'cantidad': cantidad,
            'valor_unitario': valor_unitario,
            'peso_unitario': peso_unitario,
            'order_index': order_idx,
            'tariff_group': _tariff_group_from_pieza(pieza or ''),
        }

        items.append(item)
        order_idx += 1

        if len(items) >= 200:  # Limit
            break

    # If no items from text parsing, try table extraction
    if not items:
        items = _parse_pdf_tables(data)

    # Sort by order_index
    items.sort(key=lambda x: x.get('order_index', 0))

    return items


def _llm_extract_pdf_items(text: str) -> list[dict]:
    """Extract items using LLM (Gemini) with Redis caching"""
    import hashlib
    
    # Create hash for text to use as cache key
    text_hash = hashlib.md5(text.encode()).hexdigest()
    
    # Try to get from cache first
    if llm_cache:
        try:
            cached_result = llm_cache.get_pdf_extraction(text_hash)
            if cached_result:
                print(f"✅ Cache hit for PDF extraction (hash: {text_hash[:8]})")
                return cached_result.get('items', [])
        except Exception as e:
            print(f"⚠️ Cache retrieval error: {e}")
    
    # If not in cache, perform extraction
    try:
        from pdf_extractor import _llm_extract_pdf_items as standalone_llm
        items = standalone_llm(text)
        
        # Cache the result for future use
        if llm_cache and items:
            try:
                llm_cache.cache_pdf_extraction(text_hash, {'items': items})
                print(f"✅ Cached PDF extraction result (hash: {text_hash[:8]})")
            except Exception as e:
                print(f"⚠️ Cache storage error: {e}")
        
        return items
    except Exception as exc:
        print(f"LLM extraction bridge error: {exc}")
        return []


def _fallback_extraction(text: str) -> list[dict]:
    """Fallback robust extraction when LLM fails"""
    print("Fallback extraction...")
    items = []
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    
    # Look for lines with numbers that seem like items
    for i, line in enumerate(lines):
        # Find numbers in line
        nums = re.findall(r'\d+[\.,]\d+', line)
        if len(nums) >= 2:
            # Take last 2 numbers as quantity and price
            cantidad = _to_number_any(nums[-2])
            precio = _to_number_any(nums[-1])

            if cantidad > 0 and precio > 0:
                # Extract description (text before numbers)
                ultimo_num = nums[-1]
                idx_ultimo = line.rfind(ultimo_num)
                descripcion = line[:idx_ultimo].strip()

                if len(descripcion) > 5:
                    # Try to extract NCM
                    ncm_match = re.search(r'\b(\d{6,8})\b', line)
                    pieza = ncm_match.group(1) if ncm_match else ''

                    # Infer origin from context
                    origen = 'XX'
                    upper_line = line.upper()
                    if any(palabra in upper_line for palabra in ['BRASIL', 'BRAZIL', 'BR']):
                        origen = 'BR'
                    elif any(palabra in upper_line for palabra in ['CHINA', 'CN']):
                        origen = 'CN'
                    elif any(palabra in upper_line for palabra in ['USA', 'UNITED', 'US']):
                        origen = 'US'

                    items.append({
                        'pieza': pieza,
                        'descripcion': descripcion[:200],
                        'origen': origen,
                        'cantidad': cantidad,
                        'valor_unitario': precio,
                        'peso_unitario': 0.0,
                        'order_index': i,
                        'tariff_group': _tariff_group_from_pieza(pieza)
                    })
    # If still no items, look for any line with product description
    if not items:
        for i, line in enumerate(lines):
            if len(line) > 20 and not any(word in line.lower() for word in ['total', 'subtotal', 'iva', 'tax', 'invoice', 'factura']):
                # Check if there's any number in line
                nums = re.findall(r'\d+[\.,]\d+', line)
                if nums:
                    precio = _to_number_any(nums[0])
                    descripcion = line.strip()

                    items.append({
                        'pieza': '',
                        'descripcion': descripcion[:200],
                        'origen': 'XX',
                        'cantidad': 1.0,
                        'valor_unitario': precio,
                        'peso_unitario': 0.0,
                        'order_index': i,
                        'tariff_group': ''
                    })

    print(f"Fallback extracted {len(items)} items")
    return items


def _evaluate_extraction_quality(items: list[dict]) -> dict:
    """Evaluate quality of extracted items and determine if LLM should be used"""
    if not items:
        return {'use_llm': True, 'reason': 'no_items', 'quality_score': 0}

    quality_score = 0
    total_items = len(items)
    items_with_ncm = 0
    items_with_good_desc = 0
    items_with_origin = 0

    for item in items:
        # Evaluate NCM (6-8 digits)
        ncm = str(item.get('pieza', '')).strip()
        if ncm and len(ncm) >= 6:
            items_with_ncm += 1
            quality_score += 30  # 30 points for valid NCM

        # Evaluate description (min 10 chars, not only numbers)
        desc = str(item.get('descripcion', '')).strip()
        if desc and len(desc) >= 10 and not desc.replace(' ', '').replace(',', '').replace('.', '').isdigit():
            items_with_good_desc += 1
            quality_score += 25  # 25 points for good description

        # Evaluate origin (not XX)
        origin = str(item.get('origen', '')).strip()
        if origin and origin != 'XX':
            items_with_origin += 1
            quality_score += 15  # 15 points for specific origin

    # Calculate percentages
    ncm_coverage = items_with_ncm / total_items if total_items > 0 else 0
    desc_coverage = items_with_good_desc / total_items if total_items > 0 else 0
    origin_coverage = items_with_origin / total_items if total_items > 0 else 0

    # Normalize score to 0-100
    quality_score = min(100, quality_score)

    # Decide if LLM should be used
    use_llm = False
    reason = 'high_quality'

    # Cases where LLM SHOULD be used:
    if quality_score < 50:  # Low overall score
        use_llm = True
        reason = 'low_quality_score'
    elif ncm_coverage < 0.5:  # Less than 50% have NCM
        use_llm = True
        reason = 'low_ncm_coverage'
    elif desc_coverage < 0.7:  # Less than 70% have good description
        use_llm = True
        reason = 'low_description_coverage'
    elif total_items < 2:  # Very few items detected
        use_llm = True
        reason = 'few_items'

    return {
        'use_llm': use_llm,
        'reason': reason,
        'quality_score': quality_score,
        'ncm_coverage': ncm_coverage,
        'desc_coverage': desc_coverage,
        'origin_coverage': origin_coverage,
        'total_items': total_items
    }


def _ocr_enhanced_fallback(data: bytes) -> list[dict]:
    """OCR enhanced fallback for scanned PDFs"""
    try:
        with tempfile.TemporaryDirectory() as td:
            pdf_path = os.path.join(td, 'in.pdf')
            with open(pdf_path, 'wb') as f:
                f.write(data)

            # Strategy 1: pdftoppm (most common)
            img_files = []
            try:
                img_base = os.path.join(td, 'page')
                subprocess.run(['pdftoppm', '-jpeg', '-f', '1', '-l', '3', pdf_path, img_base],
                             check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                img_files = sorted(glob.glob(os.path.join(td, 'page-*.jpg')))
            except Exception:
                pass

            # Strategy 2: pdf2image (Python alternative)
            if not img_files:
                try:
                    from pdf2image import convert_from_path
                    images = convert_from_path(pdf_path, first_page=1, last_page=3, dpi=200)
                    for i, img in enumerate(images):
                        img_path = os.path.join(td, f'page_{i}.jpg')
                        img.save(img_path, 'JPEG')
                        img_files.append(img_path)
                except Exception:
                    pass

            if not img_files:
                return []

            # OCR with multiple languages and configurations
            import pytesseract
            text_all = ''

            # OCR configurations
            configs = [
                '--psm 6 -l spa+eng+por',  # Uniform block mode, Spanish+English+Portuguese
                '--psm 3 -l spa+eng+por',  # Auto mode, multi-language
                '--psm 6 -l eng',          # English only
                '--psm 6 -l spa',          # Spanish only
            ]

            for img_path in img_files[:2]:  # Only first 2 pages
                for config in configs:
                    try:
                        text = pytesseract.image_to_string(img_path, config=config)
                        if text.strip():
                            text_all += text + '\n'
                            break  # If one config works, move to next image
                    except Exception:
                        continue

            if not text_all.strip():
                return []

            # Clean OCR text
            text_all = re.sub(r'\n\s*\n\s*\n', '\n\n', text_all)
            text_all = re.sub(r'[^\w\sÀ-ÿ.,;:()/\-]', '', text_all)

            if len(text_all.strip()) < 10:
                print(f"OCR fallback: text too short ({len(text_all.strip())} chars)")
                return []

            # Use multilayer system with OCR text
            print(f"OCR fallback: processing {len(text_all)} chars of OCR text")
            ocr_items = _llm_extract_pdf_items(text_all[:8000])

            if ocr_items:
                # Mark items as OCR source
                for item in ocr_items:
                    item['ocr_source'] = True
                print(f"OCR fallback: extracted {len(ocr_items)} items successfully")
                return ocr_items
            else:
                print(f"OCR fallback: no items extracted from OCR text")

            # Fallback: use basic line parsing
            lines = [ln.strip() for ln in text_all.splitlines() if ln.strip()]
            return _parse_ci_text_lines(lines)

    except Exception as e:
        print(f"OCR fallback error: {e}")
        return []


def _parse_ci_text_lines(lines: list[str]) -> list[dict]:
    """Parse CI-like text lines"""
    # Locate header
    header_idx = None
    blob = ' '.join(lines).lower()
    default_origin = _detect_default_origin(lines)
    for i, ln in enumerate(lines[:200]):
        low = ln.lower()
        if 'model' in low and 'price' in low and ('qty' in low or 'quantity' in low):
            header_idx = i
            break
    if header_idx is None:
        return []
    items = []
    number = re.compile(r'^-?\d+(?:[\.,]\d+)?$')
    for idx, ln in enumerate(lines[header_idx+1: header_idx+1+300], start=1):
        s = re.sub(r'\s+', ' ', ln.strip())
        if not s:
            continue
        toks = s.split(' ')
        if len(toks) < 4:
            continue
        # Find numeric tail
        def isnum(x):
            try:
                float(x.replace(',', '.'))
                return True
            except Exception:
                return False
        if not isnum(toks[-2]):
            continue
        qty = _to_float(toks[-2], 0.0)
        price = _to_float(toks[-3], 0.0) if len(toks) >= 3 and isnum(toks[-3]) else 0.0
        # Remove leading index
        start_idx = 0
        if isnum(toks[0]):
            start_idx = 1
        if len(toks) > start_idx+1 and isnum(toks[start_idx]):
            start_idx += 1
        model = ' '.join(toks[start_idx: -3 if isnum(toks[-3]) else -2]).strip()
        if not model or _is_noise_desc(model):
            continue
        if qty <= 0 and price <= 0:
            continue
        items.append({
            'pieza': '',
            'descripcion': model,
            'origen': default_origin,
            'cantidad': qty if qty > 0 else 1.0,
            'valor_unitario': price if price >= 0 else 0.0,
            'peso_unitario': 0.0,
            'order_index': idx,
            'tariff_group': '',
        })
        if len(items) >= 100:
            break
    return items


def _parse_text_tabular_lines(lines: list[str]) -> list[dict]:
    """Heuristic for tabular text without table structure"""
    items: list[dict] = []
    header_idx = -1
    has_kg = False
    # Locate header
    for i, ln in enumerate(lines[:300]):
        low = ln.lower()
        if (("descripcion" in low or "descrição" in low or "description" in low or 'item description' in low)
            and ("qty" in low or "quantity" in low or "cantidad" in low or "pcs" in low or "qty/cant" in low)
            and (("rate" in low) or ("us$/pc" in low) or ("price" in low) or ("valor" in low) or ("unit price" in low) or ("p.u." in low) or ("pu" in low))
            and ("amount" in low or "total" in low or 'importe' in low)):
            header_idx = i
            has_kg = ("kg" in low) or ("cantidad kg" in low)
            break
    if header_idx == -1:
        # Alternative: detect block with NCM and table below
        for i, ln in enumerate(lines[:300]):
            if 'ncm' in ln.lower() and re.search(r'\d{6,8}', ln):
                header_idx = i + 1
                has_kg = True
                break
    if header_idx == -1:
        return []

    current_ncm = ''
    desc_buffer = ''
    for idx, ln in enumerate(lines[header_idx+1: header_idx+1+500], start=1):
        s = ln.strip()
        if not s:
            continue
        low = s.lower()
        if any(stop in low for stop in ['observ', 'precio total', 'price total', 'comments', 'comentarios', 'subtotal', 'grand total', 'total factura', 'total invoice', 'iva', 'tax', 'impuesto']):
            break
        m_ncm = re.search(r'ncm\s*(\d{6,8})', low)
        if m_ncm:
            current_ncm = m_ncm.group(1)
        # Detect numbers at end
        nums = list(re.finditer(r'(?:\d{1,3}(?:[\.,]\d{3})*|\d+)(?:[\.,]\d{1,3})?', s))
        if len(nums) < 2:
            # Description continuation line
            if len(s) > 2:
                desc_buffer = (desc_buffer + ' ' + s).strip()
            continue
        # Take last 4 as [kg?, qty, price, total]
        last = nums[-4:] if len(nums) >= 4 else nums[-3:]
        # Start of first number marks end of description
        desc_end = last[0].start()
        part = s[:desc_end].strip(' -:\t')
        desc = (desc_buffer + ' ' + part).strip() if desc_buffer else part
        desc_buffer = ''
        if len(desc) < 3:
            continue
        vals = [ _to_number_any(s[m.start():m.end()]) for m in last ]
        vals = [v for v in vals if v is not None]
        if len(vals) < 2:
            continue
        if len(last) == 4:
            kg, qty, price, total = vals[-4], vals[-3], vals[-2], vals[-1]
        elif len(last) == 3:
            kg, qty, price, total = 0.0, vals[-3], vals[-2], vals[-1]
        else:
            continue
        if qty is None or qty <= 0:
            continue
        price = price or 0.0
        if _is_noise_desc(desc):
            continue
        items.append({
            'pieza': current_ncm or '',
            'descripcion': desc,
            'origen': 'XX',
            'cantidad': qty,
            'valor_unitario': price,
            'peso_unitario': kg or 0.0,
            'order_index': idx,
            'tariff_group': _tariff_group_from_pieza(current_ncm or ''),
        })
        if len(items) >= 200:
            break
    return items


# Format-specific extractors
def _extraer_maro_tecnica(texto: str) -> list[dict]:
    """Specific extraction for Maro/Tuper invoices with technical specs"""
    items = []
    # Look for technical specs like "32,00x1,50x1290,00"
    patron_especificaciones = r'(\d+[\.,]\d+)\s*x\s*(\d+[\.,]\d+)\s*x\s*(\d+[\.,]\d+)'
    matches = re.findall(patron_especificaciones, texto)

    for match in matches:
        cantidad, _, precio = match
        cantidad = float(cantidad.replace(',', '.'))
        precio = float(precio.replace(',', '.'))

        # Look for description near this spec
        patron_contexto = rf'([A-Za-zÀ-ÿ\s]+?)\s*{re.escape(match[0])}'
        contexto_match = re.search(patron_contexto, texto)

        descripcion = contexto_match.group(1).strip() if contexto_match else "Tubos de acero"

        items.append({
            'pieza': '',
            'descripcion': descripcion,
            'origen': 'BR',
            'cantidad': cantidad,
            'valor_unitario': precio,
            'peso_unitario': 0.0
        })

    return items


def _extraer_synergy_ingles(texto: str) -> list[dict]:
    """Specific extraction for Synergy invoices in English"""
    items = []
    lineas = texto.split('\n')
    descripcion_actual = ""

    for linea in lineas:
        linea = linea.strip()
        if not linea:
            continue

        # If line looks like product description
        if any(palabra in linea.upper() for palabra in ['HEAT EXCHANGER', 'SPARE PART', 'GASKET', 'PLATE']):
            descripcion_actual = linea
        elif descripcion_actual and re.search(r'\d+[\.,]\d+', linea):
            # Look for numbers in line (quantity and price)
            numeros = re.findall(r'\d+[\.,]\d+', linea)
            if len(numeros) >= 2:
                cantidad = float(numeros[0].replace(',', '.'))
                precio = float(numeros[1].replace(',', '.'))

                items.append({
                    'pieza': '',
                    'descripcion': descripcion_actual,
                    'origen': 'US',
                    'cantidad': cantidad,
                    'valor_unitario': precio,
                    'peso_unitario': 0.0
                })

                descripcion_actual = ""

    return items


def _extraer_con_ncm(texto: str) -> list[dict]:
    """Extraction for invoices with visible NCM"""
    items = []
    lineas = texto.split('\n')

    for linea in lineas:
        linea = linea.strip()
        if not linea:
            continue

        # Look for NCM in line
        ncm_match = re.search(r'\b(\d{6,8})\b', linea)
        if ncm_match:
            ncm = ncm_match.group(1)

            # Look for numbers (quantity, price)
            numeros = re.findall(r'\d+[\.,]\d+', linea)
            if len(numeros) >= 2:
                cantidad = float(numeros[0].replace(',', '.'))
                precio = float(numeros[-1].replace(',', '.'))

                # Extract description (text before NCM)
                descripcion = linea[:ncm_match.start()].strip()

                items.append({
                    'pieza': ncm,
                    'descripcion': descripcion,
                    'origen': 'XX',
                    'cantidad': cantidad,
                    'valor_unitario': precio,
                    'peso_unitario': 0.0
                })

    return items


def _extraer_generica(texto: str) -> list[dict]:
    """Generic extraction for any invoice"""
    items = []
    lineas = texto.split('\n')

    for linea in lineas:
        linea = linea.strip()
        if not linea or len(linea) < 10:
            continue

        # Look for numbers in line
        numeros = re.findall(r'\d+[\.,]\d+', linea)
        if len(numeros) >= 2:
            # Take last 2 numbers as quantity and price
            cantidad = float(numeros[-2].replace(',', '.'))
            precio = float(numeros[-1].replace(',', '.'))

            # Extract description (text before numbers)
            ultimo_numero = numeros[-1]
            idx_ultimo = linea.rfind(ultimo_numero)
            descripcion = linea[:idx_ultimo].strip()

            if len(descripcion) > 5:
                items.append({
                    'pieza': '',
                    'descripcion': descripcion,
                    'origen': 'XX',
                    'cantidad': cantidad,
                    'valor_unitario': precio,
                    'peso_unitario': 0.0
                })

    return items


def audit_log(user: Dict, action: str, detail: Dict) -> None:
    """Audit log helper"""
    import json
    logger.info(json.dumps({
        "action": action,
        "user": user.get("sub") if user else "anonymous",
        "roles": user.get("roles") if user else [],
        "detail": detail,
    }))


def _json_error(status_code: int, detail: str, *, include_items: bool = False) -> JSONResponse:
    content: Dict[str, Any] = {"success": False, "detail": detail}
    if include_items:
        content.setdefault("items", [])
    return JSONResponse(status_code=status_code, content=content)


# ============================================================================
# ENDPOINTS
# ============================================================================

@router.post('/process_operation/')
async def process_operation(body: dict = Body(...), user: dict = None):
    """Process import operations from payload (public endpoint, no auth required)"""
    payload_data = body.get('payload', body) if isinstance(body, dict) else body
    if hasattr(payload_data, 'dict'):
        payload_data = payload_data.dict()
    try:
        payload = OperationPayload(**payload_data)
    except ValidationError as exc:
        raise HTTPException(
            status_code=422,
            detail={"errors": exc.errors(), "detail": "Formato inválido para operación"}
        )

    logger.info(f"[process_operation] Processing operation with {len(payload.items)} items")
    if user:
        audit_log(user, "process_operation", {"items": len(payload.items)})
    try:
        operation_id = (payload.operation_id or "").strip()
        if not operation_id:
            logger.warning("[process_operation] Missing operation_id")
            return _json_error(400, "operation_id requerido")

        items = payload.items or []
        if not isinstance(items, list) or len(items) == 0:
            return _json_error(400, "No hay items para generar")

        for it in items:
            if not it.pieza:
                return _json_error(422, "Todos los items deben tener NCM (pieza)")
            if it.cantidad <= 0:
                return _json_error(422, f"cantidad inválida para {it.pieza}")
            if it.valor_unitario <= 0:
                return _json_error(422, f"valor_unitario inválido para {it.pieza}")
            if it.peso_unitario <= 0:
                return _json_error(422, f"peso_unitario inválido para {it.pieza}")

        GENERATED_DIR = os.path.join(DATA_DIR, 'generated')
        os.makedirs(GENERATED_DIR, exist_ok=True)

        import pandas as pd

        # Convert items to dicts and prepare data for AVG format
        items_data = []
        for it in items:
            item_dict = it.dict() if hasattr(it, 'dict') else it

            # Format according to exact AVG specification
            avg_row = {
                'Pieza': item_dict.get('pieza', ''),
                'Descripcion': item_dict.get('descripcion', ''),
                'Origen': item_dict.get('origen', 'XX'),
                'Peso Unitario': float(item_dict.get('peso_unitario', 0.0)),
                'Cantidad': float(item_dict.get('cantidad', 0.0)),
                'Valor Unitario': float(item_dict.get('valor_unitario', 0.0)),
                'Marca': item_dict.get('marca', ''),
                'Modelo': item_dict.get('modelo', ''),
                'Version': item_dict.get('version', ''),
                'otros': item_dict.get('otros', ''),
                'separador': item_dict.get('separador', ''),
                'ventaja': item_dict.get('ventaja', ''),
                'TOTAL': round(float(item_dict.get('cantidad', 0.0)) * float(item_dict.get('valor_unitario', 0.0)), 2)
            }
            items_data.append(avg_row)

        # Create DataFrame with ordered columns according to real AVG format
        avg_columns = [
            'Pieza', 'Descripcion', 'Origen', 'Peso Unitario',
            'Cantidad', 'Valor Unitario', 'Marca', 'Modelo',
            'Version', 'otros', 'separador', 'ventaja', 'TOTAL'
        ]

        df = pd.DataFrame(items_data, columns=avg_columns)

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"AVG_{ts}.xlsx"
        filepath = os.path.join(GENERATED_DIR, filename)

        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='AVG')

                # Apply professional formatting
                workbook = writer.book
                worksheet = writer.sheets['AVG']

                # Header formatting
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                for col_num, column_title in enumerate(avg_columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Adjust column widths
                column_widths = {
                    'A': 12,  # Pieza
                    'B': 35,  # Descripcion
                    'C': 8,   # Origen
                    'D': 12,  # Peso Unitario
                    'E': 10,  # Cantidad
                    'F': 15,  # Valor Unitario
                    'G': 15,  # Marca
                    'H': 20,  # Modelo
                    'I': 25,  # Version
                    'J': 20,  # otros
                    'K': 10,  # separador
                    'L': 20,  # ventaja
                    'M': 12   # TOTAL
                }

                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

        except Exception as excel_error:
            print(f"Excel generation error: {excel_error}")
            # Fallback to simple CSV with 13 columns
            import csv
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(avg_columns)
                for item_data in items_data:
                    w.writerow([item_data[col] for col in avg_columns])

        return {
            'success': True,
            'filename': filename,
            'download_url': f'/download/{filename}',
            'validated_items_count': len(items)
        }
    except Exception:
        logger.exception("Error generating AVG file")
        return _json_error(500, "Error generando archivo AVG")


@router.post('/upload_pdf')
@router.post('/upload_pdf/')
async def upload_pdf(file: UploadFile = File(...), user: dict = Depends(require_role("operador"))):
    """
    Extract items from invoice PDF using robust pipeline with multilayer system
    SECURITY: File validation with MIME type checking
    """
    try:
        # Security: Validate file upload (MIME type, size, extension)
        if SECURITY_MODULES_AVAILABLE:
            try:
                max_size = int(os.environ.get('MAX_UPLOAD_MB', '10')) * 1024 * 1024
                data = await validate_file_upload(file, file_type='pdf', max_size=max_size)
            except Exception as e:
                safe_msg = get_safe_error_message(e, debug=False)
                return _json_error(400, safe_msg, include_items=True)
        else:
            # Fallback: Basic validation
            max_mb = float(os.environ.get('MAX_UPLOAD_MB') or 10)
            data = await file.read()
            if data and (len(data) / (1024*1024)) > max_mb:
                raise HTTPException(
                    status_code=413,
                    detail=f'Archivo excede tamaño permitido ({max_mb} MB)'
                )
            _basic_pdf_security_checks(file, data)

        # Use robust extraction pipeline
        items = _robust_extract_pdf_items(data)

        # Evaluate extraction quality
        quality = _evaluate_extraction_quality(items)

        # Log quality assessment
        ncm_coverage = quality.get('ncm_coverage', 0) or 0
        desc_coverage = quality.get('desc_coverage', 0) or 0
        total_items = quality.get('total_items', 0) or 0
        print(f"Quality Assessment: Score={quality['quality_score']}, NCM={ncm_coverage:.1%}, Desc={desc_coverage:.1%}, Items={total_items}")

        # If quality is low or user wants LLM fallback, use ROBUST multilayer system
        if quality['use_llm'] or str(os.environ.get('FORCE_LLM_EXTRACTION') or 'false').lower() in ('1','true','yes','on'):
            print(f"Using ROBUST LLM multilayer system: {quality['reason']}")
            text = _extract_pdf_text(data)
            if text:
                # Use robust multilayer system
                llm_items = []
                try:
                    llm_items = _llm_extract_pdf_items(text[:8000])
                    print(f"LLM extracted {len(llm_items)} items")
                except Exception as llm_error:
                    print(f"LLM error: {llm_error}, using fallback...")
                    llm_items = _fallback_extraction(text)

                if llm_items:
                    items = llm_items
                    print(f"ROBUST multilayer extracted {len(llm_items)} items")
                else:
                    print("No items from multilayer, keeping original extraction")

        # Optional OCR if still no items and OCR enabled
        if not items and str(os.environ.get('ENABLE_PDF_OCR') or 'false').lower() in ('1','true','yes','on'):
            ocr_items = _ocr_enhanced_fallback(data)
            if ocr_items:
                items = ocr_items

        # Ensure stable order
        items.sort(key=lambda x: x.get('order_index', 0))

        success = len(items) > 0
        return {
            'success': success,
            'items': items,
            'extraction_method': 'llm_multilayer' if quality['use_llm'] else 'traditional',
            'quality_assessment': quality
        }
    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        return _json_error(500, str(e), include_items=True)


@router.post('/upload_pdf/public')
@router.post('/upload_pdf/public/')
async def upload_pdf_public(file: UploadFile = File(...)):
    """
    Extract items from invoice PDF using robust pipeline - PUBLIC VERSION (no auth required for demo)
    SECURITY: File validation with MIME type checking
    """
    try:
        # Security: Validate file upload (MIME type, size, extension)
        if SECURITY_MODULES_AVAILABLE:
            try:
                max_size = int(os.environ.get('MAX_UPLOAD_MB', '10')) * 1024 * 1024
                data = await validate_file_upload(file, file_type='pdf', max_size=max_size)
            except Exception as e:
                safe_msg = get_safe_error_message(e, debug=False)
                return _json_error(400, safe_msg, include_items=True)
        else:
            # Fallback: Basic validation
            max_mb = float(os.environ.get('MAX_UPLOAD_MB') or 10)
            data = await file.read()
            if data and (len(data) / (1024*1024)) > max_mb:
                raise HTTPException(
                    status_code=413,
                    detail=f'Archivo excede tamaño permitido ({max_mb} MB)'
                )
            _basic_pdf_security_checks(file, data)

        # Use robust extraction pipeline
        items = _robust_extract_pdf_items(data)

        # Evaluate extraction quality
        quality = _evaluate_extraction_quality(items)

        # Log quality assessment
        ncm_coverage = quality.get('ncm_coverage', 0) or 0
        desc_coverage = quality.get('desc_coverage', 0) or 0
        total_items = quality.get('total_items', 0) or 0
        print(f"[PUBLIC] Quality Assessment: Score={quality['quality_score']}, NCM={ncm_coverage:.1%}, Desc={desc_coverage:.1%}, Items={total_items}")

        # If quality is low or user wants LLM fallback, use ROBUST multilayer system
        if quality['use_llm'] or str(os.environ.get('FORCE_LLM_EXTRACTION') or 'false').lower() in ('1','true','yes','on'):
            print(f"[PUBLIC] Using ROBUST LLM multilayer system: {quality['reason']}")
            text = _extract_pdf_text(data)
            if text:
                # Use robust multilayer system
                llm_items = []
                try:
                    llm_items = _llm_extract_pdf_items(text[:8000])
                    print(f"[PUBLIC] LLM extracted {len(llm_items)} items")
                except Exception as llm_error:
                    print(f"[PUBLIC] LLM error: {llm_error}, using fallback...")
                    llm_items = _fallback_extraction(text)

                if llm_items:
                    items = llm_items
                    print(f"[PUBLIC] ROBUST multilayer extracted {len(llm_items)} items")
                else:
                    print("[PUBLIC] No items from multilayer, keeping original extraction")

        # Optional OCR if still no items and OCR enabled
        if not items and str(os.environ.get('ENABLE_PDF_OCR') or 'false').lower() in ('1','true','yes','on'):
            ocr_items = _ocr_enhanced_fallback(data)
            if ocr_items:
                items = ocr_items

        # Ensure stable order
        items.sort(key=lambda x: x.get('order_index', 0))

        success = len(items) > 0
        return {
            'success': success,
            'items': items,
            'extraction_method': 'llm_multilayer' if quality['use_llm'] else 'traditional',
            'quality_assessment': quality
        }
    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        return _json_error(500, str(e), include_items=True)


@router.post('/upload_pdf_llm')
@router.post('/upload_pdf_llm/')
async def upload_pdf_llm(file: UploadFile = File(...), user: dict = Depends(require_role("operador"))):
    """
    Extract items from invoice PDF using GEMINI ALWAYS ARCHITECTURE.
    Always prioritizes AI for maximum extraction quality.
    SECURITY: File validation with MIME type checking
    """
    try:
        # Security: Validate file upload (MIME type, size, extension)
        if SECURITY_MODULES_AVAILABLE:
            try:
                max_size = int(os.environ.get('MAX_UPLOAD_MB', '10')) * 1024 * 1024
                data = await validate_file_upload(file, file_type='pdf', max_size=max_size)
            except Exception as e:
                safe_msg = get_safe_error_message(e, debug=False)
                return _json_error(400, safe_msg, include_items=True)
        else:
            # Fallback: Basic size check
            max_mb = float(os.environ.get('MAX_UPLOAD_MB') or 10)
            data = await file.read()
            if data and (len(data) / (1024*1024)) > max_mb:
                return {'success': False, 'items': [], 'detail': f'Archivo excede tamaño permitido ({max_mb} MB)'}

        print("Using ROBUST LLM multilayer system")
        text = _extract_pdf_text(data)
        if not text:
            return {'success': False, 'items': [], 'detail': 'No se pudo extraer texto del PDF'}

        print(f"Text extracted: {len(text)} characters")
        print(f"Config: ENABLE_PDF_LLM_FALLBACK={os.environ.get('ENABLE_PDF_LLM_FALLBACK')}, GEMINI_API_KEY present: {'✅' if os.environ.get('GEMINI_API_KEY') else '❌'}")

        # GEMINI ALWAYS ARCHITECTURE - Always prioritize AI
        llm_items = []
        extraction_method = 'unknown'
        debug_info = {
            'text_length': len(text),
            'config_ok': bool(os.environ.get('ENABLE_PDF_LLM_FALLBACK') == 'true' and os.environ.get('GEMINI_API_KEY')),
            'strategies_attempted': [],
            'architecture': 'gemini_always'
        }

        print("GEMINI ALWAYS ARCHITECTURE: AI-First approach for maximum quality")
        print(f"Config status: LLM enabled: {debug_info['config_ok']}, Text length: {len(text)} chars")

        # STRATEGY 1: Gemini with full text
        try:
            print("Gemini 1.5 Flash: Full text extraction (8000 chars)")
            debug_info['strategies_attempted'].append('gemini_full_text')
            llm_items = _llm_extract_pdf_items(text[:8000])
            if llm_items:
                extraction_method = 'gemini_full_text'
                print(f"Gemini successful with full text: {len(llm_items)} items")
            else:
                print("Gemini with full text: no items extracted")
        except Exception as e:
            print(f"Gemini full text error: {e}")
            debug_info['strategies_attempted'].append('gemini_full_text_error')

        # STRATEGY 2: Gemini with shorter text (if previous failed)
        if not llm_items and len(text) > 4000:
            try:
                print("Gemini 1.5 Flash: Shorter text extraction (4000 chars)")
                debug_info['strategies_attempted'].append('gemini_short_text')
                llm_items = _llm_extract_pdf_items(text[:4000])
                if llm_items:
                    extraction_method = 'gemini_short_text'
                    print(f"Gemini successful with short text: {len(llm_items)} items")
                else:
                    print("Gemini with short text: no items extracted")
            except Exception as e:
                print(f"Gemini short text error: {e}")
                debug_info['strategies_attempted'].append('gemini_short_text_error')

        # STRATEGY 3: Gemini with minimal text (last AI attempt)
        if not llm_items and len(text) > 2000:
            try:
                print("Gemini 1.5 Flash: Minimal text extraction (2000 chars)")
                debug_info['strategies_attempted'].append('gemini_minimal_text')
                llm_items = _llm_extract_pdf_items(text[:2000])
                if llm_items:
                    extraction_method = 'gemini_minimal_text'
                    print(f"Gemini successful with minimal text: {len(llm_items)} items")
                else:
                    print("Gemini with minimal text: no items extracted")
            except Exception as e:
                print(f"Gemini minimal text error: {e}")
                debug_info['strategies_attempted'].append('gemini_minimal_text_error')

        # STRATEGY 4: Only if Gemini fails completely (emergency)
        if not llm_items:
            print("EMERGENCY FALLBACK: Gemini failed, using basic parser to avoid total failure")
            try:
                debug_info['strategies_attempted'].append('emergency_fallback')
                llm_items = _fallback_extraction(text)
                if llm_items:
                    extraction_method = 'emergency_fallback'
                    print(f"Emergency fallback successful: {len(llm_items)} items")
                else:
                    extraction_method = 'total_failure'
                    print("Total extraction failure - no items found")
            except Exception as e:
                print(f"Emergency fallback error: {e}")
                debug_info['strategies_attempted'].append('emergency_fallback_error')
                extraction_method = 'total_failure'

        # STRATEGY 5: Basic line extraction (if all else fails)
        if not llm_items:
            try:
                print("Strategy 5: Basic line extraction")
                debug_info['strategies_attempted'].append('basic_line_extraction')
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                for i, line in enumerate(lines[:50]):  # Only first 50 lines
                    nums = re.findall(r'\d+[\.,]\d+', line)
                    if len(nums) >= 2:
                        cantidad = _to_number_any(nums[-2])
                        precio = _to_number_any(nums[-1])
                        if cantidad > 0 and precio > 0:
                            desc = line[:line.rfind(nums[-1])].strip()
                            if len(desc) > 5:
                                llm_items.append({
                                    'pieza': '',
                                    'descripcion': desc[:200],
                                    'origen': 'XX',
                                    'cantidad': cantidad,
                                    'valor_unitario': precio,
                                    'peso_unitario': 0.0,
                                    'order_index': i,
                                    'tariff_group': ''
                                })
                if llm_items:
                    extraction_method = 'basic_line_extraction'
                    print(f"Strategy 5 successful: {len(llm_items)} items")
                else:
                    print("Strategy 5 failed: no items extracted")
            except Exception as e:
                print(f"Strategy 5 error: {e}")
                debug_info['strategies_attempted'].append('basic_line_extraction_error')

        # If still no items, create at least one with basic info
        if not llm_items:
            print("Strategy 6: Emergency item creation")
            debug_info['strategies_attempted'].append('emergency_item')
            # Look for any number that could be price
            prices = re.findall(r'\d+[\.,]\d+', text)
            if prices:
                emergency_price = _to_number_any(prices[0])
                llm_items.append({
                    'pieza': '',
                    'descripcion': 'Producto detectado en factura (extracción de emergencia)',
                    'origen': 'XX',
                    'cantidad': 1.0,
                    'valor_unitario': emergency_price,
                    'peso_unitario': 0.0,
                    'order_index': 1,
                    'tariff_group': ''
                })
                extraction_method = 'emergency_item'
                print(f"Strategy 6 successful: 1 emergency item created")
            else:
                debug_info['strategies_attempted'].append('emergency_item_failed')

        # Ensure stable order
        llm_items.sort(key=lambda x: x.get('order_index', 0))

        success = len(llm_items) > 0
        debug_info['items_extracted'] = len(llm_items)
        debug_info['extraction_method'] = extraction_method

        print(f"Final result: {len(llm_items)} items using {extraction_method}")

        # Register Gemini metrics if infrastructure available
        if INFRASTRUCTURE_AVAILABLE:
            try:
                gemini_used = any('gemini' in strategy for strategy in debug_info['strategies_attempted'])
                gemini_success = extraction_method.startswith('gemini_')

                # Register in monitoring service
                monitoring_service.record_api_request(
                    endpoint=f"pdf_extraction_{extraction_method}",
                    response_time_ms=0,
                    is_error=not success
                )

                # Gemini specific log
                if app_logger and gemini_used:
                    app_logger.info(f"Gemini extraction: method={extraction_method}, success={gemini_success}, items={len(llm_items)}",
                                  extra={
                                      "event_type": "gemini_extraction",
                                      "extraction_method": extraction_method,
                                      "items_extracted": len(llm_items),
                                      "success": gemini_success,
                                      "strategies_attempted": debug_info['strategies_attempted']
                                  })
            except Exception as e:
                print(f"Error registering Gemini metrics: {e}")

        return {
            'success': success,
            'items': llm_items,
            'extraction_method': extraction_method,
            'total_items': len(llm_items),
            'debug_info': debug_info,
            'strategies_attempted': debug_info['strategies_attempted']
        }
    except Exception as e:
        return _json_error(500, str(e), include_items=True)


@router.post('/upload_pdf_gemini_only')
@router.post('/upload_pdf_gemini_only/')
async def upload_pdf_gemini_only(file: UploadFile = File(...), user: dict = Depends(require_role("operador"))):
    """
    GEMINI ONLY: Extract items ONLY with Gemini 1.5 Flash - NO FALLBACKS.
    100% AI architecture as requested by user.
    SECURITY: File validation with MIME type checking
    """
    try:
        # Security: Validate file upload (MIME type, size, extension)
        if SECURITY_MODULES_AVAILABLE:
            try:
                max_size = int(os.environ.get('MAX_UPLOAD_MB', '10')) * 1024 * 1024
                data = await validate_file_upload(file, file_type='pdf', max_size=max_size)
            except Exception as e:
                safe_msg = get_safe_error_message(e, debug=False)
                return _json_error(400, safe_msg, include_items=True)
        else:
            # Fallback: Basic size check
            max_mb = float(os.environ.get('MAX_UPLOAD_MB') or 10)
            data = await file.read()
            if data and (len(data) / (1024*1024)) > max_mb:
                return {'success': False, 'items': [], 'detail': f'Archivo excede tamaño permitido ({max_mb} MB)'}

        print("GEMINI ONLY ENDPOINT: 100% AI extraction, no fallbacks")

        # Extract text
        text = _extract_pdf_text(data)
        if not text:
            return {'success': False, 'items': [], 'detail': 'No se pudo extraer texto del PDF'}

        print(f"Text extracted: {len(text)} characters")

        # Verify Gemini configuration
        enable_llm = str(os.environ.get('ENABLE_PDF_LLM_FALLBACK') or 'false').lower()
        api_key = os.environ.get('GEMINI_API_KEY')

        if not (enable_llm in ('1','true','yes','on') and api_key):
            return {
                'success': False,
                'items': [],
                'detail': 'Gemini no está configurado. Verifica ENABLE_PDF_LLM_FALLBACK=true y GEMINI_API_KEY'
            }

        print("Gemini configuration verified")

        # ONLY GEMINI - No fallbacks
        llm_items = []
        extraction_method = 'gemini_only_failed'

        # Try with different text sizes
        text_sizes = [
            ('full', text[:8000]),
            ('medium', text[:4000]),
            ('short', text[:2000]),
            ('minimal', text[:1000])
        ]

        for size_name, text_chunk in text_sizes:
            if llm_items:  # If we already have items, don't try more
                break

            try:
                print(f"Gemini 1.5 Flash attempt with {size_name} text ({len(text_chunk)} chars)")
                llm_items = _llm_extract_pdf_items(text_chunk)

                if llm_items:
                    extraction_method = f'gemini_only_{size_name}'
                    print(f"Gemini successful with {size_name} text: {len(llm_items)} items")
                    break
                else:
                    print(f"Gemini with {size_name} text: no items extracted")

            except Exception as e:
                print(f"Gemini {size_name} text error: {e}")
                continue

        if not llm_items:
            return {
                'success': False,
                'items': [],
                'detail': 'Gemini no pudo extraer items del PDF. Intenta con un PDF diferente o usa el endpoint con fallbacks.',
                'extraction_method': 'gemini_only_failed',
                'debug_info': {
                    'text_length': len(text),
                    'attempts_made': len(text_sizes),
                    'architecture': 'gemini_only'
                }
            }

        # Enrich extracted items
        enriched_items = []
        for item in llm_items:
            enriched_item = {
                'pieza': item.get('pieza', ''),
                'descripcion': item.get('descripcion', ''),
                'marca': item.get('marca', ''),
                'modelo': item.get('modelo', ''),
                'version': item.get('version', ''),
                'otros': item.get('otros', ''),
                'ventaja': item.get('ventaja', ''),
                'separador': '',
                'origen': item.get('origen', ''),
                'cantidad': float(item.get('cantidad', 0)),
                'valor_unitario': float(item.get('valor_unitario', 0)),
                'peso_unitario': float(item.get('peso_unitario', 0)),
                'total': float(item.get('cantidad', 0)) * float(item.get('valor_unitario', 0))
            }
            enriched_items.append(enriched_item)

        return {
            'success': True,
            'items': enriched_items,
            'extraction_method': extraction_method,
            'debug_info': {
                'text_length': len(text),
                'architecture': 'gemini_only',
                'items_extracted': len(enriched_items)
            }
        }

    except Exception as e:
        return {'success': False, 'items': [], 'detail': f'Error en Gemini Only: {str(e)}'}
