# === GENERADOR DE EXCEL EN FORMATO AVG ===
# Este módulo es responsable de crear archivos Excel con el formato
# exacto requerido por el sistema MARIA de despachantes de aduana.

import pandas as pd  # Para manipulación de datos y exportación a Excel
from typing import List, Optional  # Type hints para mejor documentación
from proyecto_maria.models.operations import Item  # Modelo Pydantic de items
from datetime import datetime  # Para timestamps únicos en nombres de archivo
from openpyxl import load_workbook  # Para estilos avanzados
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # Estilos
from openpyxl.utils import get_column_letter  # Utilidades
import os

def create_maria_excel(items: List[Item], operation_id: str, user_plan: str = 'basic',
                        user_email: Optional[str] = None, client_name: Optional[str] = None) -> str:
    """
    🏭 GENERADOR PRINCIPAL DE EXCEL AVG

    Convierte una lista de items validados en un archivo Excel con el formato
    exacto requerido por el sistema MARIA de despachantes de aduana.

    Args:
        items: Lista de objetos Item ya validados
        operation_id: Identificador único de la operación
        user_plan: Plan del usuario ('basic' o 'premium') - define watermark
        user_email: Email del usuario (opcional, para metadata)
        client_name: Nombre del cliente (opcional, para metadata Premium)

    Returns:
        str: Nombre del archivo Excel generado (ej: AVG_25DI00241_20240915_143022.xlsx)

    Proceso:
    1. Convierte items Pydantic → diccionarios Python
    2. Calcula columna TOTAL (cantidad × valor_unitario)
    3. Crea DataFrame de Pandas
    4. Renombra columnas al formato AVG exacto
    5. Reordena columnas según especificación MARIA
    6. Exporta a Excel con timestamp único
    7. Aplica estilos profesionales (colores, fuentes, bordes)
    8. Agrega watermark según plan (Basic: texto / Premium: logo)
    9. Agrega footer con metadata
    """
    if not items:
        raise ValueError("No hay ítems válidos para generar el Excel.")

    # Convertir la lista de objetos Pydantic a una lista de diccionarios
    items_data = [item.model_dump() for item in items]

    # Calcular el TOTAL (cantidad * valor_unitario)
    for item in items_data:
        item['TOTAL'] = item['cantidad'] * item['valor_unitario']

    # Crear un DataFrame de Pandas
    df = pd.DataFrame(items_data)

    # Renombrar columnas para coincidir exactamente con formato AVG
    column_mapping = {
        'pieza': 'Pieza',
        'codigo_parte': 'Cod.Parte',
        'descripcion': 'Descripcion',
        'origen': 'Origen',
        'peso_unitario': 'Peso Unitario',
        'cantidad': 'Cantidad',
        'valor_unitario': 'Valor Unitario',
        'marca': 'Marca',
        'modelo': 'Modelo',
        'version': 'Version',
        'otros': 'otros ',
        'separador': 'separador',
        'ventaja': 'ventaja ',
        'TOTAL': 'TOTAL'
    }

    df = df.rename(columns=column_mapping)

    # Definir el orden exacto de las columnas según formato AVG
    column_order = [
        'Pieza', 'Cod.Parte', 'Descripcion', 'Origen', 'Peso Unitario', 'Cantidad',
        'Valor Unitario', 'Marca', 'Modelo', 'Version', 'otros ',
        'separador', 'ventaja ', 'TOTAL'
    ]

    # Reordenar las columnas (solo las que existen)
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]

    # Generar un nombre de archivo único con formato AVG
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"AVG_{operation_id.replace(' ', '_')}_{timestamp}.xlsx"

    # Guardar en DATA_DIR (/CDI/data/) para que el endpoint /download/ pueda encontrarlo
    # Path: core/excel_generator.py -> proyecto_maria/ -> CDI/ -> CDI/data/
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data')
    os.makedirs(data_dir, exist_ok=True)
    filepath = os.path.join(data_dir, filename)

    # Exportar a Excel sin índices
    df.to_excel(filepath, index=False, engine='openpyxl')

    # ===== APLICAR ESTILOS PROFESIONALES =====
    workbook = load_workbook(filepath)
    worksheet = workbook.active

    # Aplicar estilos
    apply_professional_styles(worksheet, len(items))

    # Agregar watermark según plan
    add_watermark(worksheet, user_plan, len(items))

    # Agregar footer con metadata
    add_footer_metadata(worksheet, len(items), operation_id, user_plan, user_email, client_name)

    # Guardar cambios
    workbook.save(filepath)

    return filename


def apply_professional_styles(worksheet, num_rows):
    """
    Aplica estilos profesionales al Excel: colores corporativos, fuentes, bordes.

    Args:
        worksheet: Objeto worksheet de openpyxl
        num_rows: Número de filas de datos (sin contar header)
    """
    # Definir estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')  # Azul corporativo
    data_font = Font(name='Arial', size=10)
    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Aplicar estilos al header (fila 1)
    for col_num in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Aplicar estilos a datos (filas 2 en adelante)
    for row_num in range(2, num_rows + 2):
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Ajustar ancho de columnas automáticamente
    for col_num in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in worksheet[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Máximo 50
        worksheet.column_dimensions[column_letter].width = adjusted_width


def add_watermark(worksheet, user_plan, num_rows):
    """
    Agrega marca de agua según el plan del usuario.

    Args:
        worksheet: Objeto worksheet de openpyxl
        user_plan: 'basic' o 'premium'
        num_rows: Número de filas de datos
    """
    if user_plan == 'basic':
        # Marca de agua textual diagonal para Basic
        # Agregar texto en una celda combinada al final
        watermark_row = num_rows + 3
        watermark_cell = worksheet.cell(row=watermark_row, column=1)
        watermark_cell.value = "⚠️  VERSIÓN BÁSICA - MARÍA  ⚠️"
        watermark_cell.font = Font(name='Arial', size=14, bold=True, color='FFA500', italic=True)
        watermark_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Combinar celdas para el watermark
        worksheet.merge_cells(
            start_row=watermark_row,
            start_column=1,
            end_row=watermark_row,
            end_column=worksheet.max_column
        )

        # Fondo amarillo suave
        watermark_fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
        for col_num in range(1, worksheet.max_column + 1):
            worksheet.cell(row=watermark_row, column=col_num).fill = watermark_fill

    else:
        # Para Premium: agregar logo (si existe) o texto elegante
        watermark_row = num_rows + 3
        watermark_cell = worksheet.cell(row=watermark_row, column=1)
        watermark_cell.value = "✨ Generado con MARÍA Premium ✨"
        watermark_cell.font = Font(name='Arial', size=12, bold=True, color='FFD700')
        watermark_cell.alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells(
            start_row=watermark_row,
            start_column=1,
            end_row=watermark_row,
            end_column=worksheet.max_column
        )


def add_footer_metadata(worksheet, num_rows, operation_id, user_plan, user_email, client_name):
    """
    Agrega footer con metadata: fecha, usuario, plan, versión.

    Args:
        worksheet: Objeto worksheet de openpyxl
        num_rows: Número de filas de datos
        operation_id: ID de la operación
        user_plan: Plan del usuario
        user_email: Email del usuario (opcional)
        client_name: Nombre del cliente (opcional)
    """
    footer_row = num_rows + 5  # 2 filas debajo del watermark

    # Crear texto del footer
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    footer_parts = [
        f"Generado: {timestamp}",
        f"Operación: {operation_id}",
        f"Plan: {user_plan.upper()}"
    ]

    if user_email:
        footer_parts.append(f"Usuario: {user_email}")

    if client_name and user_plan == 'premium':
        footer_parts.append(f"Cliente: {client_name}")

    footer_parts.append("Sistema: MARÍA v2.0")

    footer_text = " | ".join(footer_parts)

    # Agregar footer
    footer_cell = worksheet.cell(row=footer_row, column=1)
    footer_cell.value = footer_text
    footer_cell.font = Font(name='Arial', size=8, color='6B7280', italic=True)
    footer_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Combinar celdas para el footer
    worksheet.merge_cells(
        start_row=footer_row,
        start_column=1,
        end_row=footer_row,
        end_column=worksheet.max_column
    )

